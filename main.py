"""
しきしまの家 営農部 LINE業務サポートボット v5
- 全メッセージ保存・さかのぼり検索
- 作業報告を自動記録・分類（Haiku）
- ToDoの自動抽出・管理
- 決定事項の自動記録
- 未決案件の検知・フォローアップ
- 知識の蓄積（「覚えておいて」）
- @メンションには何でも回答（Sonnet）
- 画像を送ると内容を解析・説明（Sonnet Vision）
- 毎朝9時：スルー検知・未決フォローアップ
- 毎週月曜：週報自動投
- Googleカレンダーに作業記録を転記
"""

import os
import json
import re
import sqlite3
import base64
from datetime import datetime, timedelta
from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage, JoinEvent, ImageMessage, FileMessage
import anthropic
from apscheduler.schedulers.background import BackgroundScheduler
from googleapiclient.discovery import build
from google.oauth2 import service_account

app = Flask(__name__)

LINE_CHANNEL_SECRET = os.environ.get('LINE_CHANNEL_SECRET', '')
LINE_CHANNEL_ACCESS_TOKEN = os.environ.get('LINE_CHANNEL_ACCESS_TOKEN', '')
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

_data_dir = '/data' if os.path.isdir('/data') else '/tmp'  # Railway永続ボリューム対応
DB_PATH = os.environ.get('DB_PATH', os.path.join(_data_dir, 'messages.db'))
GOOGLE_SERVICE_ACCOUNT_JSON = os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON', '')
GOOGLE_CALENDAR_ID = os.environ.get('GOOGLE_CALENDAR_ID', 'primary')

MODEL_FAST  = "claude-haiku-4-5-20251001"
MODEL_SMART = "claude-sonnet-4-20250514"

WORK_CATEGORIES = [
    "水稲", "大豆", "野菜", "農機・施設管理", "除草・畔草刈",
    "水管理・用水路", "環境整備", "共同作業一般", "組合業務", "研修・視察",
    "くるみ（脱穀）", "くるみ（選別）", "くるみ（その他）",
    "野菜（個人）", "米（個人）", "くるみ（個人）", "その他個人"
]

UNANSWERED_THRESHOLD_HOURS = 24
PENDING_FOLLOWUP_DAYS = 3
WEEKDAY_JP = ['月', '火', '水', '木', '金', '土', '日']

STATE_ASKING_HOURS  = 'asking_hours'
STATE_ASKING_PEOPLE = 'asking_people'
STATE_CONFIRMING    = 'confirming'

CORRECTION_KEYWORDS = [
    '修正', '違う', 'ちがう', '間違', 'いや', 'そうじゃない',
    '取り消し', '削除', '消して', '直して', '変えて', 'ちょっと待って'
]

KNOWLEDGE_KEYWORDS = ['覚えておいて', '記録しておいて', 'メモしておいて']

_BOT_USER_ID = None


# ==================== DB初期化 ====================

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS all_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            group_id TEXT NOT NULL,
            user_id TEXT, user_name TEXT,
            message TEXT,
            message_id TEXT UNIQUE,
            message_type TEXT DEFAULT 'general'
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            group_id TEXT NOT NULL,
            user_id TEXT, user_name TEXT, message TEXT,
            message_id TEXT UNIQUE,
            needs_reply INTEGER DEFAULT 0,
            replied INTEGER DEFAULT 0,
            work_category TEXT, work_hours REAL,
            work_date TEXT, work_style TEXT, raw_analysis TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS groups (
            group_id TEXT PRIMARY KEY, group_name TEXT, joined_at TEXT, scheduler_enabled INTEGER DEFAULT 1
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS conversations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id TEXT NOT NULL,
            user_id TEXT NOT NULL,
            state TEXT NOT NULL,
            partial_analysis TEXT,
            original_message TEXT,
            last_event_id TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS todos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id TEXT NOT NULL,
            assignee TEXT,
            task TEXT NOT NULL,
            source_message TEXT,
            created_by TEXT,
            created_at TEXT,
            due_date TEXT,
            status TEXT DEFAULT 'open',
            completed_at TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS decisions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id TEXT NOT NULL,
            decision TEXT NOT NULL,
            source_message TEXT,
            decided_by TEXT,
            created_at TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS pending_issues (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id TEXT NOT NULL,
            summary TEXT NOT NULL,
            source_message TEXT,
            raised_by TEXT,
            created_at TEXT,
            last_followup_at TEXT,
            status TEXT DEFAULT 'open',
            resolved_at TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS knowledge (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id TEXT NOT NULL,
            content TEXT NOT NULL,
            stored_by TEXT,
            created_at TEXT,
            tag TEXT,
            scope TEXT DEFAULT 'group'
        )
    ''')
    # 既存DBにscopeカラムがない場合は追加（マイグレーション）
    try:
        c.execute("ALTER TABLE knowledge ADD COLUMN scope TEXT DEFAULT 'group'")
        conn.commit()
    except Exception:
        pass  # すでに存在する場合はスキップ
    c.execute('''
        CREATE TABLE IF NOT EXISTS recent_media (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id TEXT NOT NULL,
            user_id TEXT,
            message_id TEXT NOT NULL,
            filename TEXT,
            media_type TEXT,
            timestamp TEXT NOT NULL,
            file_content TEXT
        )
    ''')
    # マイグレーション: file_content カラムが無い場合に追加
    try:
        c.execute('ALTER TABLE recent_media ADD COLUMN file_content TEXT')
        conn.commit()
    except Exception:
        pass  # すでに存在する場合は無視
    conn.commit()
    conn.close()


init_db()


def save_recent_media(group_id, user_id, message_id, filename, media_type, file_content=None):
    """ファイル・画像のメッセージIDをDBに保存（自動分析しない）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.execute(
        'INSERT INTO recent_media (group_id, user_id, message_id, filename, media_type, timestamp, file_content) VALUES (?,?,?,?,?,?,?)',
        (group_id, user_id, message_id, filename, media_type, ts, file_content)
    )
    conn.commit()
    conn.close()


def update_file_content(group_id, message_id, file_content):
    """ファイル分析後にテキスト内容をDBに保存（Fix1: 次の会話でも参照可能に）"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        'UPDATE recent_media SET file_content=? WHERE group_id=? AND message_id=?',
        (file_content, group_id, message_id)
    )
    conn.commit()
    conn.close()


def get_recent_media(group_id, media_type=None, minutes=60):
    """直近N分以内の最新ファイル/画像情報を取得"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    cutoff = (datetime.now() - timedelta(minutes=minutes)).strftime('%Y-%m-%d %H:%M:%S')
    if media_type:
        c.execute(
            'SELECT message_id, filename, media_type, user_id, file_content FROM recent_media WHERE group_id=? AND media_type=? AND timestamp>=? ORDER BY timestamp DESC LIMIT 1',
            (group_id, media_type, cutoff)
        )
    else:
        c.execute(
            'SELECT message_id, filename, media_type, user_id, file_content FROM recent_media WHERE group_id=? AND timestamp>=? ORDER BY timestamp DESC LIMIT 1',
            (group_id, cutoff)
        )
    row = c.fetchone()
    conn.close()
    return row  # (message_id, filename, media_type, user_id, file_content)


def get_recent_file_content(group_id, minutes=1440):
    """最近分析済みのファイルテキストを取得（デフォルト24時間以内）。次の会話でも参照できるように。"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    cutoff = (datetime.now() - timedelta(minutes=minutes)).strftime('%Y-%m-%d %H:%M:%S')
    c.execute(
        'SELECT filename, file_content FROM recent_media WHERE group_id=? AND media_type=? AND file_content IS NOT NULL AND timestamp>=? ORDER BY timestamp DESC LIMIT 1',
        (group_id, 'file', cutoff)
    )
    row = c.fetchone()
    conn.close()
    return row  # (filename, file_content) or None


# ==================== スケジューラ ====================

def daily_check():
    """毎朝9時：未返信チェック + 未決案件フォローアップ"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT group_id FROM groups WHERE scheduler_enabled = 1')
    groups = [row[0] for row in c.fetchall()]
    conn.close()
    for gid in groups:
        notify_unanswered(gid)
        followup_pending_issues(gid)


def weekly_report_job():
    """毎週月曜9時：週報を全グループに投稿"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT group_id FROM groups WHERE scheduler_enabled = 1')
    groups = [row[0] for row in c.fetchall()]
    conn.close()
    for gid in groups:
        report = build_weekly_report(gid)
        if report:
            line_bot_api.push_message(gid, TextSendMessage(text=report))


scheduler = BackgroundScheduler(timezone='Asia/Tokyo')
scheduler.add_job(daily_check, 'cron', hour=9, minute=0)
scheduler.add_job(weekly_report_job, 'cron', day_of_week='mon', hour=9, minute=0)
scheduler.start()


# ==================== Webhook ====================

@app.route("/callback", methods=['POST'])
def callback():
    signature = request.headers.get('X-Line-Signature', '')
    body = request.get_data(as_text=True)
    try:
        payload = json.loads(body)
        for ev in payload.get('events', []):
            msg_type = ev.get('message', {}).get('type', '')
            ev_type = ev.get('type', '')
            print(f"[WEBHOOK] ev={ev_type} msg={msg_type}", flush=True)
            if ev_type == 'message' and msg_type == 'file':
                handle_file_raw(ev)
    except Exception as e:
        print(f"[WEBHOOK ERR] {e}", flush=True)
    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)
    except Exception as e:
        print(f"[HANDLER ERR] {e}", flush=True)
    return 'OK'


# ==================== グループ参加 ====================

@handler.add(JoinEvent)
def handle_join(event):
    if event.source.type == 'group':
        group_id = event.source.group_id
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            'INSERT OR IGNORE INTO groups (group_id, joined_at) VALUES (?, ?)',
            (group_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        )
        conn.commit()
        conn.close()
        line_bot_api.reply_message(event.reply_token, TextSendMessage(
            text=(
                "オルオルです🦉\n\n"
                "太古から農村を見守ってきた者として、\n"
                "皆さんのお役に立てれば幸いです。\n\n"
                "@オルオル と呼びかけてください。\n"
                "どんな質問・相談にも答えます。\n\n"
                "【コマンド】\n"
                "「/未返信」→ 返信待ちメッセージ一覧\n"
                "「/タスク」→ 未完了タスク一覧\n"
                "「/決定事項」→ 最近の決定事項\n"
                "「/未決」→ 検討中案件一覧\n"
                "「/知識」→ このグループ専用の知識🔒\n"
                "「/共通知識」→ 全グループ共通の知識🌐\n"
                "「/ヘルプ」→ 使い方"
            )
        ))


# ==================== メンション検知 ====================

def get_bot_user_id():
    global _BOT_USER_ID
    if not _BOT_USER_ID:
        try:
            _BOT_USER_ID = line_bot_api.get_bot_info().user_id
        except Exception as e:
            print(f"get_bot_info error: {e}")
    return _BOT_USER_ID


def is_bot_mentioned(event):
    # 自然な「オルオル」呼びかけを検知（@メンション不要）
    BOT_NAMES = ['オルオル', 'おるおる']
    text = getattr(event.message, 'text', '') or ''
    if any(name in text for name in BOT_NAMES):
        return True
    # 公式@メンションも検知
    try:
        msg = event.message
        if hasattr(msg, 'mention') and msg.mention:
            bot_id = get_bot_user_id()
            if bot_id:
                for m in msg.mention.mentionees:
                    if hasattr(m, 'user_id') and m.user_id == bot_id:
                        return True
    except Exception as e:
        print(f"mention check error: {e}")
    return False

def extract_mention_text(text):
    # @メンションを除去
    text = re.sub(r'@\S+\s*', '', text).strip()
    # 「オルオル、」「オルオル！」などの自然な呼びかけを先頭から除去
    text = re.sub(r'^[オお][ルる][オお][ルる][、。！？!?,\s]*', '', text).strip()
    return text


def get_user_name(group_id, user_id):
    try:
        profile = line_bot_api.get_group_member_profile(group_id, user_id)
        return profile.display_name
    except Exception:
        return user_id


# ==================== 画像メッセージ対応（Sonnet Vision） ====================

@handler.add(MessageEvent, message=ImageMessage)
def handle_image(event):
    if event.source.type != 'group':
        return
    group_id = event.source.group_id
    user_id  = event.source.user_id
    # 画像IDを保存するだけ（自動分析しない）
    # 「@オルオル この画像見て」と言われた時に分析する
    save_recent_media(group_id, user_id, event.message.id, '画像', 'image')


# ==================== ファイル対応（txt/pdf/docx/xlsx/csv） ====================

def extract_file_text(file_bytes, filename):
    """ファイル種別に応じてテキストを抽出する"""
    import io, csv
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    try:
        if ext == 'txt':
            try:
                return file_bytes.decode('utf-8')
            except UnicodeDecodeError:
                return file_bytes.decode('shift_jis', errors='replace')
        elif ext == 'pdf':
            import pdfplumber
            parts = []
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        parts.append(t)
            return '\n'.join(parts) if parts else '（PDFからテキストを抽出できませんでした）'
        elif ext == 'docx':
            from docx import Document
            doc = Document(io.BytesIO(file_bytes))
            return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
        elif ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f'【シート: {sheet_name}】')
                for row in ws.iter_rows(values_only=True):
                    row_text = '\t'.join(str(c) if c is not None else '' for c in row)
                    if row_text.strip():
                        parts.append(row_text)
            return '\n'.join(parts)
        elif ext == 'csv':
            try:
                text = file_bytes.decode('utf-8')
            except UnicodeDecodeError:
                text = file_bytes.decode('shift_jis', errors='replace')
            reader = csv.reader(io.StringIO(text))
            return '\n'.join('\t'.join(row) for row in reader)
        else:
            return f'（未対応の形式です: .{ext}）'
    except Exception as e:
        return f'（読み込みエラー: {e}）'


def handle_file_raw(ev):
    """ファイルイベントのメッセージIDをDBに保存するだけ（自動分析しない）"""
    source = ev.get('source', {})
    src_type = source.get('type', '')
    if src_type not in ('group', 'room'):
        return
    group_id   = source.get('groupId') or source.get('roomId', '')
    user_id    = source.get('userId', '')
    msg        = ev.get('message', {})
    filename   = msg.get('fileName', 'unknown')
    message_id = msg.get('id', '')
    save_recent_media(group_id, user_id, message_id, filename, 'file')
    print(f"[FILE] saved: group={group_id} file={filename}", flush=True)
    # 返信しない（スタッフ同士のファイル共有を邪魔しない）


@handler.add(MessageEvent, message=FileMessage)
def handle_file(event):
    """ファイルをDBに静かに保存するだけ。返信しない。
    @オルオルで依頼されたときだけ分析・返答する。"""
    if event.source.type != 'group':
        return
    group_id  = event.source.group_id
    user_id   = event.source.user_id
    filename  = event.message.file_name
    message_id = event.message.id

    # まずDBにメタ情報を保存（reply_tokenを消費しないよう返信しない）
    save_recent_media(group_id, user_id, message_id, filename, 'file')
    print(f"[FILE] received: group={group_id} file={filename}", flush=True)

    # LINEのファイルダウンロードトークンは期限があるので、
    # 今のうちにテキスト抽出してDBに保存しておく
    try:
        message_content = line_bot_api.get_message_content(message_id)
        file_bytes = b''.join(chunk for chunk in message_content.iter_content())
        file_text = extract_file_text(file_bytes, filename)
        if file_text.strip():
            MAX_CHARS = 8000
            if len(file_text) > MAX_CHARS:
                file_text = file_text[:MAX_CHARS]
            update_file_content(group_id, message_id, file_text)
            print(f"[FILE] text saved to DB: {filename} ({len(file_text)} chars)", flush=True)
    except Exception as e:
        print(f"[FILE] download/extract error (silent): {e}", flush=True)
    # 返信しない — スタッフ間のファイル共有を邪魔しない




def analyze_image_with_sonnet(image_base64, user_name, user_instruction=None):
    if not ANTHROPIC_API_KEY:
        return "申し訳ありません、AIサービスに接続できません。"
    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        # Fix2: ユーザーの具体的な指示があれば使う
        question = user_instruction if user_instruction else "これは何ですか？詳しく教えてください。"
        response = client.messages.create(
            model=MODEL_SMART,
            max_tokens=600,
            system=(
                "あなたは「オルオル」。日本の豊かな農村に太古の昔から棲み続けるフクロウの神様です。\n"
                "送られてきた画像を見て、農業・自然・生き物・植物・病害虫・農作物・農機具・"
                "現場の状況など、何でも詳しく解説してください。\n"
                "【回答ルール】\n"
                "- 何が写っているか簡潔に特定する\n"
                "- 農業や現場作業に関連する情報があれば詳しく補足する\n"
                "- 病害虫・雑草の場合は対処法も添える\n"
                "- LINEなので長くても15行以内\n"
                "- 絵文字は🦉を中心に適度に使う\n"
                "- 判別が難しい場合や画像が不鮮明な場合は正直にその旨を伝える\n"
                "- 見えていないもの・判断できないことは断言しない"
            ),
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_base64}},
                    {"type": "text", "text": f"{user_name}さんの質問: {question}"}
                ]
            }]
        )
        return response.content[0].text.strip()
    except Exception as e:
        print(f"Image analysis error: {e}")
        return f"⚠️ 画像の解析中にエラーが発生しました: {str(e)[:50]}"


# ==================== メッセージ受信（メイン） ====================

@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    if event.source.type != 'group':
        return

    group_id     = event.source.group_id
    user_id      = event.source.user_id
    message_text = event.message.text.strip()
    message_id   = event.message.id
    timestamp    = datetime.fromtimestamp(event.timestamp / 1000).strftime('%Y-%m-%d %H:%M:%S')
    user_name    = get_user_name(group_id, user_id)

    # 全メッセージを保存
    save_all_message(timestamp, group_id, user_id, user_name, message_text, message_id)

    # ① コマンド
    if message_text.startswith('/'):
        handle_command(event, message_text, group_id)
        return

    # ② @メンション
    if is_bot_mentioned(event):
        handle_mention(event, message_text, user_name, group_id)
        return

    # ③ 「覚えておいて」系キーワード検知（文脈で判断）
    if any(kw in message_text for kw in KNOWLEDGE_KEYWORDS):
        # キーワードを除いて内容があれば記録（「覚えておいて！」だけは無視）
        stripped = message_text
        for kw in KNOWLEDGE_KEYWORDS:
            stripped = stripped.replace(kw, '').replace('：', '').replace(':', '')
        if stripped.strip().rstrip('！!。、？?').strip():
            handle_knowledge_store(event, message_text, user_name, group_id)
            return

    # ④ 会話ステート
    pending = get_pending_state(group_id, user_id)
    if pending:
        handle_conversation_response(
            event, pending, message_text, user_name,
            group_id, user_id, timestamp, message_id
        )
        return

    # ⑤ 通常メッセージ解析（Haiku）
    analysis = analyze_message_full(message_text, user_name)
    mark_replied_context(group_id, user_id, message_text)

    save_message(timestamp, group_id, user_id, user_name,
                 message_text, message_id, analysis, create_calendar=False)

    process_metadata(analysis, message_text, user_name, group_id, timestamp)

    if not analysis.get('work_category'):
        return

    process_work_report(event, analysis, message_text, user_name, group_id, user_id)


# ==================== 全メッセージ保存 ====================

def save_all_message(timestamp, group_id, user_id, user_name, message, message_id, message_type='general'):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('''
            INSERT OR IGNORE INTO all_messages
            (timestamp, group_id, user_id, user_name, message, message_id, message_type)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (timestamp, group_id, user_id, user_name, message, message_id, message_type))
        conn.commit()
    except Exception as e:
        print(f"save_all_message error: {e}")
    finally:
        conn.close()


# ==================== ToDo・決定事項・未決案件の処理 ====================

def process_metadata(analysis, message_text, user_name, group_id, timestamp):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    if analysis.get('has_todo') and analysis.get('todo_text'):
        c.execute('''
            INSERT INTO todos (group_id, assignee, task, source_message, created_by, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (group_id, analysis.get('todo_assignee'), analysis['todo_text'],
              message_text, user_name, timestamp))

    if analysis.get('has_decision') and analysis.get('decision_text'):
        c.execute('''
            INSERT INTO decisions (group_id, decision, source_message, decided_by, created_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (group_id, analysis['decision_text'], message_text, user_name, timestamp))

    if analysis.get('is_pending_issue') and analysis.get('pending_summary'):
        c.execute('''
            INSERT INTO pending_issues (group_id, summary, source_message, raised_by, created_at)
            VALUES (?, ?, ?, ?, ?)
        ''', (group_id, analysis['pending_summary'], message_text, user_name, timestamp))

    conn.commit()
    conn.close()


# ==================== 知識ベース ====================

SHARED_KNOWLEDGE_KEYWORDS = ['共通で', '全グループで', 'みんなで覚えて', '共通知識として', '全体で覚えて']

def handle_knowledge_store(event, message_text, user_name, group_id):
    content = message_text

    # scopeの判定（「共通で覚えておいて」→ shared）
    scope = 'group'
    for sk in SHARED_KNOWLEDGE_KEYWORDS:
        if sk in content:
            scope = 'shared'
            content = content.replace(sk, '').strip()
            break

    for kw in KNOWLEDGE_KEYWORDS:
        content = content.replace(kw, '').strip()
    content = content.lstrip('：:').strip()

    if not content:
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="覚えておきたい内容を「覚えておいて：〇〇」の形で教えてください🦉\n全グループ共通で覚えたい場合は「共通で覚えておいて：〇〇」")
        )
        return

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        'INSERT INTO knowledge (group_id, content, stored_by, created_at, scope) VALUES (?, ?, ?, ?, ?)',
        (group_id, content, user_name, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), scope)
    )
    conn.commit()
    conn.close()

    if scope == 'shared':
        scope_label = "【全グループ共通】"
        cmd_hint = "「/共通知識」で確認できます。"
    else:
        scope_label = "【このグループ専用】"
        cmd_hint = "「/知識」で確認できます。"

    line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(text=f"🦉 覚えました！{scope_label}\n「{content[:50]}」\n\n{cmd_hint}")
    )


# ==================== @メンション対応（Sonnet） ====================

def get_thinking_message(query):
    """クエリの内容に応じた自然な「考え中」メッセージを返す"""
    import random
    q = query
    if any(kw in q for kw in ['タスク', 'やること', 'TODO', 'todo', '未完了', 'やってない', 'やり残し']):
        return random.choice(['タスクを確認しますね🦉', 'ToDoリスト、見てみますね🦉'])
    elif any(kw in q for kw in ['スケジュール', '予定', 'カレンダー', 'いつ', '何時', '何日']):
        return random.choice(['スケジュールを調べますね🦉', '予定を確認しますね🦉'])
    elif any(kw in q for kw in ['決定', '決まった', 'ルール', '方針', '決めた']):
        return '決定事項を確認しますね🦉'
    elif any(kw in q for kw in ['未決', '保留', '宿題', 'ペンディング', 'スルー', '忘れ']):
        return random.choice(['保留案件を調べますね🦉', 'スルーがないか確認しますね🦉'])
    elif any(kw in q for kw in ['書いて', '作って', '考えて', '提案', '案を', '文章', 'メール', '返信']):
        return random.choice(['考えてみますね🦉', 'ちょっと考えてみます🦉'])
    elif any(kw in q for kw in ['まとめ', '整理', '要点', 'サマリー', 'ポイント']):
        return 'まとめてみますね🦉'
    elif any(kw in q for kw in ['先週', '先月', '過去', '前に', 'この間', 'この前', 'いつだっけ']):
        return '過去のやり取りを確認しますね🦉'
    elif any(kw in q for kw in ['誰', 'だれ', '誰が', '誰に', '何を', '何が', 'どこ', 'なぜ', 'なんで']):
        return random.choice(['調べますね🦉', 'ちょっと確認しますね🦉'])
    elif '?' in q or '？' in q:
        return random.choice(['うーん🦉', 'ちょっと待って🦉', '確認しますね🦉'])
    else:
        return random.choice(['少し考えますね🦉', 'ちょっと待ってね🦉', 'うーん🦉'])

def handle_mention(event, message_text, user_name, group_id):
    query = extract_mention_text(message_text)
    if not query:
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text="はい！何でもお気軽にどうぞ🦉")
        )
        return

    # ファイル・画像分析リクエストを検知
    FILE_KEYWORDS = ['読んで', '読み込んで', 'ファイル', '分析して', '要約して', 'まとめて', '内容見て']
    IMAGE_KEYWORDS = ['画像', '写真', '見て', '解析して', 'これどう', '何これ']
    is_file_req  = any(kw in query for kw in FILE_KEYWORDS)
    is_image_req = any(kw in query for kw in IMAGE_KEYWORDS)

    if is_file_req:
        recent = get_recent_media(group_id, 'file')
        if recent:
            message_id, filename, _, _, saved_content = recent
            # すでにDB保存済みのテキストがあればそれを使う（再ダウンロード不要）
            if saved_content:
                file_text = saved_content
                print(f"[FILE] using cached content for {filename}", flush=True)
                # キャッシュ使用時は reply_token を使って「考えますね」と返す
                try:
                    line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text=f'「{filename}」の内容を確認しますね🦉…')
                    )
                except Exception:
                    pass
            else:
                try:
                    line_bot_api.reply_message(
                        event.reply_token,
                        TextSendMessage(text=f'「{filename}」を読み込みますね🦉…')
                    )
                except Exception:
                    pass
                try:
                    message_content = line_bot_api.get_message_content(message_id)
                    file_bytes = b''.join(chunk for chunk in message_content.iter_content())
                except Exception as e:
                    line_bot_api.push_message(group_id, TextSendMessage(text=f'ファイルの取得に失敗しました🦉（{e}）'))
                    return
                file_text = extract_file_text(file_bytes, filename)
                if not file_text.strip():
                    line_bot_api.push_message(group_id, TextSendMessage(text=f'「{filename}」からテキストを読み取れませんでした🦉'))
                    return
                # Fix1: 抽出したテキストをDBに保存（次の会話でも参照できるように）
                update_file_content(group_id, message_id, file_text[:8000])

            MAX_CHARS = 8000
            truncated_note = ''
            if len(file_text) > MAX_CHARS:
                file_text = file_text[:MAX_CHARS]
                truncated_note = f'\n\n※長いため最初の{MAX_CHARS}文字のみ読み込みました。'

            # Fix2: ユーザーの実際の指示をプロンプトに使う（固定の「要約」から脱却）
            user_instruction = query
            for kw in FILE_KEYWORDS:
                user_instruction = user_instruction.replace(kw, '')
            user_instruction = user_instruction.strip().rstrip('。、！？!?,').strip()
            if not user_instruction:
                user_instruction = '内容を確認して要点をまとめてください'

            prompt = f'【ファイル「{filename}」の内容】\n{file_text}\n\n【{user_name}さんの指示】\n{user_instruction}'
            context  = get_group_context(group_id)
            history  = search_history(filename, group_id)
            knowledge = get_knowledge_context(group_id)
            reply = ask_sonnet(prompt, user_name, context, history, knowledge)
            try:
                line_bot_api.push_message(group_id, TextSendMessage(text=reply + truncated_note))
            except Exception as e:
                print(f"[PUSH ERROR] {e}", flush=True)
            return
        else:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text='直近1時間以内に送られたファイルが見つかりません🦉\nファイルを送ってから「@オルオル 読んで」と言ってね！')
            )
            return

    if is_image_req:
        recent = get_recent_media(group_id, 'image')
        if recent:
            message_id, filename, _, _, _ = recent
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text='画像を確認しますね🦉…')
            )
            try:
                message_content = line_bot_api.get_message_content(message_id)
                image_data = b''.join(chunk for chunk in message_content.iter_content())
                image_base64 = base64.standard_b64encode(image_data).decode('utf-8')
            except Exception as e:
                line_bot_api.push_message(group_id, TextSendMessage(text=f'画像の取得に失敗しました🦉（{e}）'))
                return
            # Fix2: ユーザーの具体的な質問を画像分析に渡す
            user_instruction = query
            for kw in IMAGE_KEYWORDS:
                user_instruction = user_instruction.replace(kw, '')
            user_instruction = user_instruction.strip().rstrip('。、！？!?,').strip()
            reply = analyze_image_with_sonnet(image_base64, user_name, user_instruction or None)
            try:
                line_bot_api.push_message(group_id, TextSendMessage(text=reply))
            except Exception as e:
                print(f"[PUSH ERROR] {e}", flush=True)
            return
        else:
            line_bot_api.reply_message(
                event.reply_token,
                TextSendMessage(text='直近1時間以内に送られた画像が見つかりません🦉\n画像を送ってから「@オルオル 見て」と言ってね！')
            )
            return

    # 通常のテキスト質問
    # Fix1: 最近分析したファイル内容があれば文脈に含める（ファイルについての追加質問に対応）
    file_context = ""
    recent_file = get_recent_file_content(group_id, minutes=1440)
    if recent_file:
        fname, fcontent = recent_file
        file_context = f"\n\n【直近で分析したファイル「{fname}」の内容（抜粋）】\n{fcontent[:2000]}"

    line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(text=get_thinking_message(query))
    )
    context  = get_group_context(group_id) + file_context
    history  = search_history(query, group_id)
    knowledge = get_knowledge_context(group_id)
    reply = ask_sonnet(query, user_name, context, history, knowledge)
    try:
        line_bot_api.push_message(group_id, TextSendMessage(text=reply))
    except Exception as e:
        print(f"[PUSH ERROR] {e}", flush=True)


def search_history(query, group_id, limit=20):
    """過去のメッセージからクエリに関連するものを検索"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        keywords = [w for w in re.sub(r'[^\w]', ' ', query).split() if len(w) > 1]
        results = []
        for kw in keywords[:3]:
            c.execute('''
                SELECT timestamp, user_name, message FROM all_messages
                WHERE group_id=? AND message LIKE ?
                ORDER BY timestamp DESC LIMIT 10
            ''', (group_id, f'%{kw}%'))
            results.extend(c.fetchall())
        conn.close()
        if not results:
            return ""
        seen = set()
        unique = []
        for row in sorted(results, key=lambda x: x[0], reverse=True):
            if row[2] not in seen:
                seen.add(row[2])
                unique.append(row)
        lines = ["\n【過去のやり取り（関連）】"]
        for ts, name, msg in unique[:limit]:
            lines.append(f"  {ts[:10]} {name}:「{msg[:60]}」")
        return "\n".join(lines)
    except Exception as e:
        print(f"search_history error: {e}")
        return ""


def get_knowledge_context(group_id):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        # このグループ専用 + 全グループ共通の両方を取得
        c.execute('''
            SELECT content, stored_by, created_at, scope FROM knowledge
            WHERE (group_id=? AND scope='group') OR scope='shared'
            ORDER BY scope ASC, created_at DESC LIMIT 30
        ''', (group_id,))
        rows = c.fetchall()
        conn.close()
        if not rows:
            return ""
        lines = ["\n【覚えている情報】"]
        for content, stored_by, created_at, scope in rows:
            label = "🌐共通" if scope == 'shared' else "🔒専用"
            lines.append(f"  ・[{label}] {content}（{stored_by}、{created_at[:10]}）")
        return "\n".join(lines)
    except Exception as e:
        print(f"get_knowledge_context error: {e}")
        return ""

def get_group_context(group_id):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        today = datetime.now()

        c.execute('''
            SELECT work_category, user_name, SUM(work_hours), COUNT(*)
            FROM messages
            WHERE group_id=? AND work_category IS NOT NULL
              AND work_date BETWEEN ? AND ?
            GROUP BY work_category, user_name
        ''', (group_id, today.strftime('%Y-%m-01'), today.strftime('%Y-%m-%d')))
        work_rows = c.fetchall()

        threshold = (today - timedelta(hours=UNANSWERED_THRESHOLD_HOURS)).strftime('%Y-%m-%d %H:%M:%S')
        c.execute('''
            SELECT user_name, message, timestamp FROM messages
            WHERE group_id=? AND needs_reply=1 AND replied=0 AND timestamp<?
            ORDER BY timestamp DESC LIMIT 5
        ''', (group_id, threshold))
        unanswered_rows = c.fetchall()

        c.execute('''
            SELECT assignee, task, created_at FROM todos
            WHERE group_id=? AND status='open'
            ORDER BY created_at DESC LIMIT 10
        ''', (group_id,))
        todo_rows = c.fetchall()

        c.execute('''
            SELECT summary, raised_by, created_at FROM pending_issues
            WHERE group_id=? AND status='open'
            ORDER BY created_at DESC LIMIT 5
        ''', (group_id,))
        pending_rows = c.fetchall()

        conn.close()

        lines = [f"【今日: {today.strftime('%Y年%m月%d日')}】"]

        if work_rows:
            lines.append(f"\n今月（{today.month}月）の作業記録:")
            for cat, name, hours, count in work_rows:
                h = hours or 0
                lines.append(f"  {name} / {cat}: {h:.1f}h（{count}件）")

        if unanswered_rows:
            lines.append(f"\n{UNANSWERED_THRESHOLD_HOURS}時間以上返信待ち:")
            for name, msg, ts in unanswered_rows:
                short = msg[:40] + "..." if len(msg) > 40 else msg
                lines.append(f"  {name}（{ts[:10]}）:「{short}」")

        if todo_rows:
            lines.append("\n未完了タスク:")
            for assignee, task, created_at in todo_rows:
                a = f"{assignee}さん" if assignee else "担当未定"
                lines.append(f"  ・{a}: {task}（{created_at[:10]}）")

        if pending_rows:
            lines.append("\n検討中・未決案件:")
            for summary, raised_by, created_at in pending_rows:
                lines.append(f"  ・{summary}（{raised_by}、{created_at[:10]}）")

        return "\n".join(lines)
    except Exception as e:
        print(f"get_group_context error: {e}")
        return f"（データ取得エラー: {e}）"


def ask_sonnet(query, user_name, db_context, history="", knowledge=""):
    if not ANTHROPIC_API_KEY:
        return "申し訳ありません、AIサービスに接続できません。"
    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        system_prompt = f"""あなたは「オルオル」。日本の豊かな農村に太古の昔から棲み続けるフクロウの神様です。
夜の生態系の頂点に立ち、里山が豊かである限りそこに宿る存在として、
人々の農の営みと暮らしをずっと見守ってきました。

その長い時の中で、農業・自然・気象・生き物・人の暮らし・歴史・文化・
地域のこと、あらゆる知識と知恵を深く蓄えています。

あるときは専門家のように正確な知識で答え、
あるときは物事の本質へと導く神のように語り、
あるときは寄り添う友人のように温かく接し、
基本的にはグループメンバーを支える秘書として振る舞います。

特定の組織や団体に属する存在ではなく、
豊かな農村があればどこにでも宿る、普遍的な存在です。

{db_context}
{history}
{knowledge}

【回答ルール】
- LINEのメッセージなので短く端的に（長くても15行以内）
- 過去のやり取りを聞かれた場合は、上記の履歴から該当する内容を探して答える
- 「誰が言った」「いつ決まった」など事実確認は履歴から正確に答える
- わからない・記録がない場合は正直に「記録が見つかりません」と伝える
- 絵文字は🦉を中心に適度に使って親しみやすく
- 回答の最後に追加で確認できることがあれば一言添える

【ハリュシネーション防止ルール（最重要）】
- 自分が実際に処理・分析したファイルや画像のみ「見た」「読んだ」と言う
- 処理していないコンテンツについては「確認できていません」「ファイルを送って @オルオル 読んで と言ってください」と正直に伝える
- 不確かな情報は「〜かもしれません」「確認が必要です」と明示する
- 記録や履歴にない事実について断言しない（「〜だったと思います」ではなく「記録にはありません」）
- 「はい、見えています」「はい、読みました」などの嘘の確認をしない"""

        response = client.messages.create(
            model=MODEL_SMART,
            max_tokens=600,
            system=system_prompt,
            messages=[{"role": "user", "content": f"{user_name}さん: {query}"}]
        )
        return response.content[0].text.strip()
    except Exception as e:
        print(f"Sonnet API error: {e}")
        return f"⚠️ エラーが発生しました: {str(e)[:50]}"


# ==================== 作業報告の処理 ====================

def process_work_report(event, analysis, message_text, user_name, group_id, user_id):
    missing = get_missing_info(analysis, message_text)
    if missing:
        question = build_question(missing['state'], analysis, user_name)
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=question))
        save_pending_state(group_id, user_id, missing['state'], analysis, message_text)
    else:
        event_id = add_to_calendar(analysis, user_name, message_text)
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=build_confirmation(analysis, user_name))
        )
        save_pending_state(group_id, user_id, STATE_CONFIRMING, analysis, message_text, event_id)


def get_missing_info(analysis, message_text):
    if not analysis.get('work_hours'):
        return {'state': STATE_ASKING_HOURS}
    if not analysis.get('work_style') and _has_people_hint(message_text):
        return {'state': STATE_ASKING_PEOPLE}
    return None


def _has_people_hint(text):
    hints = ['一緒', '2人', '二人', 'と一緒', 'たちで', 'みんな', '手伝', '協力', '複数']
    return any(h in text for h in hints)


# ==================== 会話の続き ====================

def handle_conversation_response(
    event, pending, message_text, user_name,
    group_id, user_id, timestamp, message_id
):
    state   = pending['state']
    partial = pending['partial_analysis']
    is_correction = any(kw in message_text for kw in CORRECTION_KEYWORDS)

    if is_correction:
        handle_correction(event, pending, message_text, user_name, group_id, user_id)
        return

    if state == STATE_CONFIRMING:
        clear_pending_state(group_id, user_id)
        analysis = analyze_message_full(message_text, user_name)
        save_message(timestamp, group_id, user_id, user_name,
                     message_text, message_id, analysis, create_calendar=False)
        mark_replied_context(group_id, user_id, message_text)
        if analysis.get('work_category'):
            process_work_report(event, analysis, message_text, user_name, group_id, user_id)
        return

    if state == STATE_ASKING_HOURS:
        updated = parse_hours_from_text(message_text, partial)
        if not updated.get('work_hours'):
            line_bot_api.reply_message(event.reply_token, TextSendMessage(
                text="⏱️ 時間が読み取れませんでした。\n「3時間」「2.5h」「半日」などで教えてください🙏"
            ))
            return
        _finalize_or_ask_more(
            event, updated, pending['original_message'],
            user_name, group_id, user_id, timestamp, message_id
        )

    elif state == STATE_ASKING_PEOPLE:
        updated = parse_people_from_text(message_text, partial, user_name)
        event_id = add_to_calendar(updated, user_name, pending['original_message'])
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=build_confirmation(updated, user_name))
        )
        save_pending_state(
            group_id, user_id, STATE_CONFIRMING,
            updated, pending['original_message'], event_id
        )


def _finalize_or_ask_more(
    event, analysis, original_message, user_name,
    group_id, user_id, timestamp, message_id
):
    missing = get_missing_info(analysis, original_message)
    if missing:
        question = build_question(missing['state'], analysis, user_name)
        line_bot_api.reply_message(event.reply_token, TextSendMessage(text=question))
        save_pending_state(group_id, user_id, missing['state'], analysis, original_message)
    else:
        event_id = add_to_calendar(analysis, user_name, original_message)
        line_bot_api.reply_message(
            event.reply_token,
            TextSendMessage(text=build_confirmation(analysis, user_name))
        )
        save_pending_state(
            group_id, user_id, STATE_CONFIRMING,
            analysis, original_message, event_id
        )


def handle_correction(event, pending, message_text, user_name, group_id, user_id):
    last_event_id = pending.get('last_event_id')
    last_analysis = pending.get('partial_analysis', {})
    if last_event_id:
        delete_calendar_event(last_event_id)
    corrected = last_analysis.copy()
    if ANTHROPIC_API_KEY:
        try:
            client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
            cat_str = "、".join(WORK_CATEGORIES)
            prompt = (
                f"前回の作業報告と記録内容、修正指示から、正しい情報をJSONで返してください。\n\n"
                f"前回の報告: {pending.get('original_message', '')}\n"
                f"記録された内容: {json.dumps(last_analysis, ensure_ascii=False)}\n"
                f"修正指示: {message_text}\n\n"
                f"以下のJSONのみを返してください（説明文不要）:\n"
                f'{{\n  "work_category": "{cat_str} のいずれか",\n'
                f'  "work_hours": 数値またはnull,\n  "work_date": "YYYY-MM-DD",\n'
                f'  "work_style": "本田+荻原" または "本田（単独）" または "荻原（単独）" または null\n}}'
            )
            response = client.messages.create(
                model=MODEL_FAST, max_tokens=200,
                messages=[{"role": "user", "content": prompt}]
            )
            result = response.content[0].text.strip()
            if '{' in result:
                result = result[result.index('{'):result.rindex('}') + 1]
            corrected = {**last_analysis, **json.loads(result)}
        except Exception as e:
            print(f"Correction parse error: {e}")
    event_id = add_to_calendar(corrected, user_name, pending.get('original_message', ''))
    reply = "✏️ 修正しました！\n" + build_confirmation(corrected, user_name, header="")
    line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply))
    save_pending_state(
        group_id, user_id, STATE_CONFIRMING,
        corrected, pending.get('original_message', ''), event_id
    )


# ==================== メッセージ生成 ====================

def build_question(state, analysis, user_name):
    cat = analysis.get('work_category', '作業')
    if state == STATE_ASKING_HOURS:
        return (
            f"📝 {user_name}さん、{cat}の作業ですね！\n"
            f"作業時間を教えてください🙏\n"
            f"（例：「3時間」「2.5h」「半日」「終日」）"
        )
    elif state == STATE_ASKING_PEOPLE:
        return (
            f"📝 {cat}の作業ですね！\n"
            f"何人で作業しましたか？\n"
            f"① {user_name}さん単独\n"
            f"② 本田さんと2人\n"
            f"③ 荻原さんと2人\n"
            f"④ 本田さん＋荻原さんと3人"
        )
    return "詳しく教えてください。"


def build_confirmation(analysis, user_name, header="✅ 記録しました！\n"):
    cat   = analysis.get('work_category', '')
    date  = analysis.get('work_date') or datetime.now().strftime('%Y-%m-%d')
    hours = analysis.get('work_hours')
    style = analysis.get('work_style')
    try:
        d = datetime.strptime(date, '%Y-%m-%d')
        date_str = f"{d.month}/{d.day}（{WEEKDAY_JP[d.weekday()]}）"
    except Exception:
        date_str = date
    lines = [header + "━━━━━━━━━━━━"]
    lines.append(f"📋 {cat}")
    lines.append(f"📅 {date_str}")
    if hours:
        lines.append(f"⏱️ {hours:.1f}時間")
    lines.append(f"👤 {user_name}")
    if style:
        lines.append(f"👥 {style}")
    lines.append("━━━━━━━━━━━━")
    lines.append("間違いがあれば「修正して」と教えてください🙏")
    return "\n".join(lines)


# ==================== 応答パース ====================

def parse_hours_from_text(text, partial_analysis):
    updated = dict(partial_analysis)
    m = re.search(r'(\d+(?:\.\d+)?)\s*[hｈ時間]', text)
    if m:
        updated['work_hours'] = float(m.group(1))
        return updated
    if '午前' in text or '午後' in text:
        updated['work_hours'] = 3.0
    elif '半日' in text:
        updated['work_hours'] = 4.0
    elif '終日' in text or '一日' in text or '1日' in text:
        updated['work_hours'] = 8.0
    elif ANTHROPIC_API_KEY:
        try:
            client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
            r = client.messages.create(
                model=MODEL_FAST, max_tokens=30,
                messages=[{"role": "user", "content": f"「{text}」から作業時間を数値（時間単位）で抽出してください。数値のみ返して。不明なら「null」。"}]
            )
            h = r.content[0].text.strip()
            if h != 'null':
                updated['work_hours'] = float(h)
        except Exception:
            pass
    return updated


def parse_people_from_text(text, partial_analysis, user_name):
    updated = dict(partial_analysis)
    if '①' in text or '単独' in text or '一人' in text or '1人' in text:
        updated['work_style'] = None
    elif '④' in text or ('本田' in text and '荻原' in text) or '3人' in text or '三人' in text:
        updated['work_style'] = '本田+荻原'
    elif '②' in text or '本田' in text:
        updated['work_style'] = '本田（単独）'
    elif '③' in text or '荻原' in text:
        updated['work_style'] = '荻原（単独）'
    else:
        updated['work_style'] = None
    return updated


# ==================== 会話ステート管理 ====================

def get_pending_state(group_id, user_id):
    cutoff = (datetime.now() - timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT state, partial_analysis, original_message, last_event_id
        FROM conversations
        WHERE group_id=? AND user_id=? AND updated_at>?
        ORDER BY updated_at DESC LIMIT 1
    ''', (group_id, user_id, cutoff))
    row = c.fetchone()
    conn.close()
    if row:
        return {
            'state': row[0],
            'partial_analysis': json.loads(row[1]) if row[1] else {},
            'original_message': row[2] or '',
            'last_event_id': row[3]
        }
    return None


def save_pending_state(
    group_id, user_id, state, partial_analysis, original_message, last_event_id=None
):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DELETE FROM conversations WHERE group_id=? AND user_id=?', (group_id, user_id))
    c.execute('''
        INSERT INTO conversations
        (group_id, user_id, state, partial_analysis, original_message,
         last_event_id, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        group_id, user_id, state,
        json.dumps(partial_analysis, ensure_ascii=False),
        original_message, last_event_id, now, now
    ))
    conn.commit()
    conn.close()


def clear_pending_state(group_id, user_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('DELETE FROM conversations WHERE group_id=? AND user_id=?', (group_id, user_id))
    conn.commit()
    conn.close()


# ==================== コマンド ====================

def handle_command(event, text, group_id):
    cmd = text.lower().strip()
    if cmd in ['/集計', '/summary']:
        reply = get_monthly_summary(group_id)
    elif cmd in ['/未返信', '/unanswered']:
        reply = get_unanswered_list(group_id)
    elif cmd in ['/タスク', '/todo']:
        reply = get_todo_list(group_id)
    elif cmd in ['/決定事項', '/decisions']:
        reply = get_decisions_list(group_id)
    elif cmd in ['/未決', '/pending']:
        reply = get_pending_list(group_id)
    elif cmd in ['/知識', '/knowledge']:
    elif cmd in ['/共通知識', '/shared']:
    elif cmd in ['/グループid', '/groupid', '/group_id']:
        reply = f"🦉 このグループのID\n\n{group_id}\n\nRailwayの環境変数 ADMIN_GROUP_ID にこの値を設定すると、このグループが管理グループになります。"
        reply = get_shared_knowledge_list()
        reply = get_knowledge_list(group_id)
    elif cmd in ['/週報', '/weekly']:
        reply = build_weekly_report(group_id) or "📊 今週はまだデータがありません。"
    elif cmd in ['/スケジューラoff', '/scheduler off', '/scheduleroff']:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('UPDATE groups SET scheduler_enabled = 0 WHERE group_id = ?', (group_id,))
        conn.commit()
        conn.close()
        reply = 'このグループへの朝・週次レポートを停止しました🦉'
    elif cmd in ['/スケジューラon', '/scheduler on', '/scheduleron']:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('UPDATE groups SET scheduler_enabled = 1 WHERE group_id = ?', (group_id,))
        conn.commit()
        conn.close()
        reply = 'このグループへの朝・週次レポートを再開します🦉'
    elif cmd in ['/ヘルプ', '/help']:
        reply = (
            "【コマンド一覧】\n"
            "「/未返信」→ 返信待ちメッセージ\n"
            "「/タスク」→ 未完了タスク\n"
            "「/決定事項」→ 最近の決定事項\n"
            "「/未決」→ 検討中案件\n"
            "「/知識」→ このグループ専用の知識🔒\n"
            "「/共通知識」→ 全グループ共通の知識🌐\n"
            "「/週報」→ 今週のサマリー\n\n"
            "【@オルオル で何でもOK】\n"
            "質問・調べもの・過去の話…気軽に🦉\n\n"
            "【知識の記憶】\n"
            "「覚えておいて：○○」→ このグループ専用🔒\n"
            "「共通で覚えておいて：○○」→ 全グループ共通🌐\n\n"
            "【営農グループ限定】\n"
            "「/集計」→ 今月の作業時間集計\n"
            "作業報告は自動で記録・分類されます✏️"
        )
    else:
        return
    line_bot_api.reply_message(event.reply_token, TextSendMessage(text=reply))


def get_todo_list(group_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT assignee, task, created_at FROM todos
        WHERE group_id=? AND status='open'
        ORDER BY created_at DESC LIMIT 15
    ''', (group_id,))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return "✅ 未完了タスクはありません！"
    lines = [f"📋 未完了タスク（{len(rows)}件）\n"]
    for assignee, task, created_at in rows:
        a = f"{assignee}さん" if assignee else "担当未定"
        lines.append(f"・{a}: {task}\n  （{created_at[:10]}）")
    return "\n".join(lines)


def get_decisions_list(group_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    cutoff = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
    c.execute('''
        SELECT decision, decided_by, created_at FROM decisions
        WHERE group_id=? AND created_at>?
        ORDER BY created_at DESC LIMIT 15
    ''', (group_id, cutoff))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return "📝 直近30日の決定事項はまだ記録されていません。"
    lines = ["📝 最近の決定事項\n"]
    for decision, decided_by, created_at in rows:
        lines.append(f"・{decision}\n  （{decided_by}、{created_at[:10]}）")
    return "\n".join(lines)


def get_pending_list(group_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT summary, raised_by, created_at FROM pending_issues
        WHERE group_id=? AND status='open'
        ORDER BY created_at DESC LIMIT 10
    ''', (group_id,))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return "✅ 未決の案件はありません！"
    lines = [f"🤔 検討中・未決案件（{len(rows)}件）\n"]
    for summary, raised_by, created_at in rows:
        lines.append(f"・{summary}\n  （{raised_by}、{created_at[:10]}）")
    return "\n".join(lines)


def get_knowledge_list(group_id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # このグループ専用のみ表示
    c.execute('''
        SELECT content, stored_by, created_at FROM knowledge
        WHERE group_id=? AND scope='group' ORDER BY created_at DESC LIMIT 20
    ''', (group_id,))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return "🦉 このグループ専用の知識はまだありません。\n「覚えておいて：〇〇」で教えてください！\n\n共通知識は「/共通知識」で確認できます。"
    lines = [f"🔒 このグループの知識（{len(rows)}件）\n"]
    for content, stored_by, created_at in rows:
        lines.append(f"・{content}\n  （{stored_by}、{created_at[:10]}）")
    return "\n".join(lines)


def get_shared_knowledge_list():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 全グループ共通知識を表示
    c.execute('''
        SELECT content, stored_by, created_at FROM knowledge
        WHERE scope='shared' ORDER BY created_at DESC LIMIT 30
    ''')
    rows = c.fetchall()
    conn.close()
    if not rows:
        return "🌐 共通知識はまだありません。\n「共通で覚えておいて：〇〇」で登録できます！"
    lines = [f"🌐 全グループ共通の知識（{len(rows)}件）\n"]
    for content, stored_by, created_at in rows:
        lines.append(f"・{content}\n  （{stored_by}、{created_at[:10]}）")
    return "\n".join(lines)

def get_monthly_summary(group_id):
    today = datetime.now()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT work_category, SUM(work_hours), COUNT(*)
        FROM messages
        WHERE group_id=? AND work_category IS NOT NULL
          AND work_date BETWEEN ? AND ?
        GROUP BY work_category ORDER BY SUM(work_hours) DESC
    ''', (group_id, today.strftime('%Y-%m-01'), today.strftime('%Y-%m-%d')))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return f"📊 {today.month}月の作業記録はまだありません。"
    lines = [f"📊 {today.month}月の作業時間集計\n"]
    total = 0
    for cat, hours, count in rows:
        h = hours or 0
        total += h
        lines.append(f"  {cat}: {h:.1f}h（{count}件）")
    lines.append(f"\n合計: {total:.1f}h")
    return "\n".join(lines)


def get_unanswered_list(group_id):
    threshold = (
        datetime.now() - timedelta(hours=UNANSWERED_THRESHOLD_HOURS)
    ).strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT user_name, message, timestamp FROM messages
        WHERE group_id=? AND needs_reply=1 AND replied=0 AND timestamp<?
        ORDER BY timestamp DESC LIMIT 10
    ''', (group_id, threshold))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return "✅ 返信待ちのメッセージはありません！"
    lines = [f"⚠️ {UNANSWERED_THRESHOLD_HOURS}時間以上返信待ち:\n"]
    for name, msg, ts in rows:
        short = msg[:40] + "..." if len(msg) > 40 else msg
        lines.append(f"👤 {name}（{ts[:10]}）\n   「{short}」")
    return "\n\n".join(lines)


# ==================== 週報 ====================

def build_weekly_report(group_id):
    today = datetime.now()
    week_ago = (today - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute('''
        SELECT work_category, user_name, SUM(work_hours), COUNT(*)
        FROM messages
        WHERE group_id=? AND work_category IS NOT NULL AND timestamp>?
        GROUP BY work_category, user_name
    ''', (group_id, week_ago))
    work_rows = c.fetchall()

    c.execute('''
        SELECT decision, decided_by FROM decisions
        WHERE group_id=? AND created_at>? ORDER BY created_at DESC LIMIT 5
    ''', (group_id, week_ago))
    decision_rows = c.fetchall()

    c.execute('''
        SELECT assignee, task FROM todos
        WHERE group_id=? AND status='open' ORDER BY created_at LIMIT 5
    ''', (group_id,))
    todo_rows = c.fetchall()

    c.execute('''
        SELECT summary FROM pending_issues
        WHERE group_id=? AND status='open' ORDER BY created_at LIMIT 5
    ''', (group_id,))
    pending_rows = c.fetchall()

    conn.close()

    if not work_rows and not decision_rows and not todo_rows:
        return None

    lines = [f"🦉 週報（{today.strftime('%m/%d')}）\n━━━━━━━━━━━━"]

    if work_rows:
        lines.append("【今週の作業】")
        for cat, name, hours, count in work_rows:
            h = hours or 0
            lines.append(f"  {name}: {cat} {h:.1f}h")

    if decision_rows:
        lines.append("\n【今週の決定事項】")
        for decision, decided_by in decision_rows:
            lines.append(f"  ・{decision}")

    if todo_rows:
        lines.append("\n【未完了タスク】")
        for assignee, task in todo_rows:
            a = f"{assignee}さん" if assignee else "担当未定"
            lines.append(f"  ・{a}: {task}")

    if pending_rows:
        lines.append("\n【検討中の案件】")
        for (summary,) in pending_rows:
            lines.append(f"  ・{summary}")

    lines.append("━━━━━━━━━━━━")
    return "\n".join(lines)


# ==================== スルー検知・未決フォローアップ ====================

def notify_unanswered(group_id):
    msg = get_unanswered_list(group_id)
    if "返信待ち" in msg and "✅" not in msg:
        line_bot_api.push_message(group_id, TextSendMessage(text=msg))


def followup_pending_issues(group_id):
    """PENDING_FOLLOWUP_DAYS日以上経過した未決案件をフォローアップ"""
    cutoff = (datetime.now() - timedelta(days=PENDING_FOLLOWUP_DAYS)).strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT id, summary, raised_by, created_at FROM pending_issues
        WHERE group_id=? AND status='open'
          AND (last_followup_at IS NULL OR last_followup_at<?)
          AND created_at<?
        ORDER BY created_at ASC LIMIT 3
    ''', (group_id, cutoff, cutoff))
    rows = c.fetchall()

    if rows:
        lines = ["🦉 フォローアップです。以下の案件、その後いかがでしょう？\n"]
        ids = []
        for issue_id, summary, raised_by, created_at in rows:
            lines.append(f"・{summary}（{raised_by}、{created_at[:10]}）")
            ids.append(issue_id)
        line_bot_api.push_message(group_id, TextSendMessage(text="\n".join(lines)))
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for issue_id in ids:
            c.execute('UPDATE pending_issues SET last_followup_at=? WHERE id=?', (now, issue_id))
        conn.commit()

    conn.close()


def mark_replied_context(group_id, user_id, message_text):
    reply_keywords = ['確認', '了解', 'わかりました', 'ありがとう', 'ok', 'OK', '👍', '✅']
    if any(kw in message_text for kw in reply_keywords):
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            UPDATE messages SET replied=1
            WHERE group_id=? AND needs_reply=1 AND replied=0
            AND id IN (
                SELECT id FROM messages
                WHERE group_id=? AND needs_reply=1 AND replied=0
                ORDER BY timestamp DESC LIMIT 5
            )
        ''', (group_id, group_id))
        conn.commit()
        conn.close()


# ==================== Claude解析（Haiku） ====================

def analyze_message_full(text, user_name):
    """メッセージを総合解析（作業報告・ToDo・決定事項・未決案件）"""
    if not ANTHROPIC_API_KEY:
        return _simple_analyze(text)
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    categories_str = "、".join(WORK_CATEGORIES)
    today_str = datetime.now().strftime('%Y-%m-%d')
    prompt = f"""農業グループのLINEメッセージを分析し、以下のJSONのみを返してください（説明文不要）。

送信者: {user_name}
メッセージ: {text}
今日: {today_str}

{{
  "needs_reply": true/false,
  "work_category": "{categories_str} のいずれか、または null",
  "work_hours": 数値またはnull,
  "work_date": "YYYY-MM-DD" または null,
  "work_style": "本田+荻原" / "本田（単独）" / "荻原（単独）" / null,
  "has_todo": true/false,
  "todo_text": "タスクの内容" または null,
  "todo_assignee": "担当者名" または null,
  "has_decision": true/false,
  "decision_text": "決定事項の内容" または null,
  "is_pending_issue": true/false,
  "pending_summary": "検討中案件の要約" または null
}}

判断基準:
- has_todo: 「〜しておいて」「〜お願い」「〜やっておく」など具体的なタスクがある
- has_decision: 「〜することになった」「〜に決まった」「〜でいこう」など決定を示す
- is_pending_issue: 「〜どうする？」「〜検討しよう」「〜考えないと」など未解決の課題がある"""
    try:
        response = client.messages.create(
            model=MODEL_FAST, max_tokens=400,
            messages=[{"role": "user", "content": prompt}]
        )
        result_text = response.content[0].text.strip()
        if '{' in result_text:
            result_text = result_text[result_text.index('{'):result_text.rindex('}') + 1]
        return json.loads(result_text)
    except Exception as e:
        print(f"Claude API error: {e}")
        return _simple_analyze(text)


def _simple_analyze(text):
    needs_reply = any(kw in text for kw in ['？', '?', 'どうします', 'お願い', '確認', '教えて'])
    m = re.search(r'(\d+(?:\.\d+)?)\s*[hｈ時間]', text)
    work_hours = float(m.group(1)) if m else None
    work_cat = None
    for cat, kws in {
        '水稲': ['水稲', '田植え', '稲刈り', 'コンバイン'],
        '大豆': ['大豆', '枝豆'],
        '除草・畔草刈': ['除草', '草刈', '畔'],
        '水管理・用水路': ['水管理', '水路'],
        '農機・施設管理': ['トラクター', '農機', '機械'],
    }.items():
        if any(kw in text for kw in kws):
            work_cat = cat
            break
    return {
        "needs_reply": needs_reply,
        "work_category": work_cat,
        "work_hours": work_hours,
        "work_date": datetime.now().strftime('%Y-%m-%d'),
        "work_style": None,
        "has_todo": False, "todo_text": None, "todo_assignee": None,
        "has_decision": False, "decision_text": None,
        "is_pending_issue": False, "pending_summary": None
    }


# ==================== Googleカレンダー ====================

def get_calendar_service():
    if not GOOGLE_SERVICE_ACCOUNT_JSON:
        return None
    try:
        creds = service_account.Credentials.from_service_account_info(
            json.loads(GOOGLE_SERVICE_ACCOUNT_JSON),
            scopes=['https://www.googleapis.com/auth/calendar']
        )
        return build('calendar', 'v3', credentials=creds)
    except Exception as e:
        print(f"Calendar service error: {e}")
        return None


def add_to_calendar(analysis, user_name, message_text):
    if not analysis.get('work_category'):
        return None
    service = get_calendar_service()
    if not service:
        return None
    work_date = analysis.get('work_date') or datetime.now().strftime('%Y-%m-%d')
    try:
        datetime.strptime(work_date, '%Y-%m-%d')
    except ValueError:
        work_date = datetime.now().strftime('%Y-%m-%d')
    hours = analysis.get('work_hours')
    style = analysis.get('work_style')
    title_parts = [f"【{analysis['work_category']}】"]
    if hours:
        title_parts.append(f"{hours:.1f}h")
    title_parts.append(f"- {user_name}")
    if style:
        title_parts.append(f"({style})")
    event = {
        'summary': " ".join(title_parts),
        'description': f"📱 LINEからの作業報告\n👤 {user_name}\n💬 {message_text}" + (f"\n👥 {style}" if style else ""),
        'start': {'date': work_date},
        'end':   {'date': work_date},
        'colorId': '2',
    }
    try:
        result = service.events().insert(calendarId=GOOGLE_CALENDAR_ID, body=event).execute()
        return result.get('id')
    except Exception as e:
        print(f"Calendar insert error: {e}")
        return None


def delete_calendar_event(event_id):
    service = get_calendar_service()
    if not service or not event_id:
        return
    try:
        service.events().delete(calendarId=GOOGLE_CALENDAR_ID, eventId=event_id).execute()
    except Exception as e:
        print(f"Calendar delete error: {e}")


# ==================== DB保存 ====================

def save_message(
    timestamp, group_id, user_id, user_name, message, message_id,
    analysis, create_calendar=True
):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('''
            INSERT OR IGNORE INTO messages
            (timestamp, group_id, user_id, user_name, message, message_id,
             needs_reply, work_category, work_hours, work_date, work_style, raw_analysis)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            timestamp, group_id, user_id, user_name, message, message_id,
            1 if analysis.get('needs_reply') else 0,
            analysis.get('work_category'), analysis.get('work_hours'),
            analysis.get('work_date'), analysis.get('work_style'),
            json.dumps(analysis, ensure_ascii=False)
        ))
        conn.commit()
    except Exception as e:
        print(f"DB save error: {e}")
    finally:
        conn.close()
    if create_calendar and analysis.get('work_category'):
        add_to_calendar(analysis, user_name, message)


# ==================== ヘルスチェック ====================

@app.route("/health", methods=['GET'])
def health():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM all_messages')
    all_count = c.fetchone()[0]
    c.execute('SELECT COUNT(*) FROM todos WHERE status="open"')
    todo_count = c.fetchone()[0]
    c.execute('SELECT COUNT(*) FROM pending_issues WHERE status="open"')
    pending_count = c.fetchone()[0]
    conn.close()
    return {
        'status': 'ok',
        'all_messages': all_count,
        'open_todos': todo_count,
        'pending_issues': pending_count,
        'models': {'fast': MODEL_FAST, 'smart': MODEL_SMART}
    }


if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
 
